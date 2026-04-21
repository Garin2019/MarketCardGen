"""Сервис экспорта карточки товара в Excel и CSV."""
import csv
import io
import json

import openpyxl
from logger import get_logger

log = get_logger("export_service")
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from models import Product


EXPORT_HEADERS = [
    ("Артикул", "article"),
    ("Краткое описание", "short_description"),
    ("Описание", "long_description"),
    ("Категория", "category"),
    ("Ключевые слова", "keywords"),
]


def _build_row(product: Product) -> dict:
    """Формируем словарь значений из продукта."""
    kw = json.loads(product.keywords or "[]")
    return {
        "article":           product.article or "",
        "short_description": product.short_description or "",
        "long_description":  product.long_description  or "",
        "category":          product.category           or "",
        "keywords":          ", ".join(kw),
    }


# ── Excel ─────────────────────────────────────────────────────────────────────

_HEADER_FILL  = PatternFill("solid", fgColor="4F46E5")
_HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
_BORDER_SIDE  = Side(style="thin", color="D0D5E8")
_CELL_BORDER  = Border(
    left=_BORDER_SIDE, right=_BORDER_SIDE,
    top=_BORDER_SIDE,  bottom=_BORDER_SIDE,
)
_WRAP = Alignment(wrap_text=True, vertical="top")


def build_excel(product: Product) -> bytes:
    """Генерирует .xlsx и возвращает байты файла."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Карточка товара"

    headers = EXPORT_HEADERS
    row_data = _build_row(product)

    # Заголовки
    for col_idx, (label, _) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font      = _HEADER_FONT
        cell.fill      = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = _CELL_BORDER
    ws.row_dimensions[1].height = 22

    # Данные
    for col_idx, (_, key) in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=row_data[key])
        cell.alignment = _WRAP
        cell.border    = _CELL_BORDER

    # Ширина столбцов
    col_widths = {
        "article": 18, "short_description": 40,
        "long_description": 60, "category": 22,
        "keywords": 35,
    }
    for col_idx, (_, key) in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(key, 20)

    # Высота строки с данными — авто не работает в openpyxl, ставим разумный max
    long = row_data.get("long_description", "")
    ws.row_dimensions[2].height = min(max(len(long) // 60 * 15, 30), 300)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    data = buf.read()
    log.info("Excel сформирован: product_id=%d size=%d байт", product.id, len(data))
    return data


# ── CSV ───────────────────────────────────────────────────────────────────────

def build_csv(product: Product) -> str:
    """Генерирует .csv и возвращает строку (UTF-8 с BOM для Excel)."""
    headers  = EXPORT_HEADERS
    row_data = _build_row(product)

    buf = io.StringIO()
    # BOM нужен чтобы Excel корректно открывал UTF-8
    buf.write("\ufeff")
    writer = csv.writer(buf, quoting=csv.QUOTE_ALL)
    writer.writerow([label for label, _ in headers])
    writer.writerow([row_data[key] for _, key in headers])
    return buf.getvalue()
